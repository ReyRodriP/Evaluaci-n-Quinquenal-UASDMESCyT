import io
from datetime import datetime
from math import ceil

from django.contrib.auth import get_user_model
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.permissions import departamentos_permitidos
from auditoria.models import Auditoria
from evidence.models import Observacion
from organization.models import PerfilUsuario

Usuario = get_user_model()

FILTRO_POR_DEFECTO = {
    'fecha': 30,
}


def _formato_fecha(valor):
    if not valor:
        return ''
    return valor.strftime('%d/%m/%Y %H:%M')


def _formato_fecha_corta(valor):
    if not valor:
        return ''
    return valor.strftime('%d/%m/%Y')


def _base_queryset_observaciones(request):
    qs = Observacion.objects.filter(activo=True).select_related(
        'usuario',
        'version__evidencia__asignacion__indicador',
        'version__evidencia__asignacion__departamento',
        'version__evidencia__asignacion__periodo',
    )

    deptos_ids = departamentos_permitidos(request)
    if deptos_ids is not None:
        qs = qs.filter(version__evidencia__asignacion__departamento_id__in=deptos_ids)

    periodo = request.query_params.get('periodo')
    if periodo:
        qs = qs.filter(version__evidencia__asignacion__periodo_id=periodo)

    departamento = request.query_params.get('departamento')
    if departamento:
        qs = qs.filter(version__evidencia__asignacion__departamento_id=departamento)

    usuario = request.query_params.get('usuario')
    if usuario:
        qs = qs.filter(usuario_id=usuario)

    return qs


def _filas_observaciones(qs):
    conteos = dict(
        qs.values_list('version__evidencia_id')
        .annotate(total=Count('id'))
        .order_by()
    )

    filas = []
    for obs in qs:
        asignacion = obs.version.evidencia.asignacion
        filas.append([
            obs.version.evidencia.titulo,
            asignacion.indicador.nombre,
            asignacion.departamento.nombre,
            asignacion.periodo.nombre,
            obs.version.version,
            obs.usuario.username if obs.usuario else '',
            obs.comentario,
            _formato_fecha(obs.fecha_creacion),
            conteos.get(obs.version.evidencia_id, 0),
        ])

    return filas


def _filas_auditoria(qs):
    filas = []
    for registro in qs:
        filas.append([
            registro.usuario.username if registro.usuario else '',
            registro.accion,
            registro.modelo,
            registro.registro_id or '',
            registro.descripcion,
            _formato_fecha(registro.fecha),
        ])
    return filas


def _filas_usuarios(qs):
    filas = []
    for usuario in qs:
        perfil = getattr(usuario, 'perfilusuario', None)
        filas.append([
            usuario.username,
            f"{usuario.first_name} {usuario.last_name}".strip(),
            usuario.email,
            (usuario.groups.first().name if usuario.groups.exists() else ''),
            perfil.departamento.nombre if perfil and perfil.departamento else '',
            _formato_fecha(usuario.last_login),
            'Activo' if usuario.is_active else 'Inactivo',
        ])
    return filas


def _paginar(request, filas):
    try:
        page = max(int(request.query_params.get('page', 1)), 1)
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
    except (TypeError, ValueError):
        page_size = 10

    total = len(filas)
    inicio = (page - 1) * page_size
    return filas[inicio:inicio + page_size], total, page, page_size, ceil(total / page_size) if total else 0


def _build_pdf(titulo, columnas, filas):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=titulo,
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = estilos['Title']
    estilo_titulo.fontSize = 14
    estilo_titulo.spaceAfter = 2

    estilo_subtitulo = estilos['Normal']
    estilo_subtitulo.fontSize = 9
    estilo_subtitulo.alignment = 1
    estilo_subtitulo.spaceAfter = 10

    estilo_celda = estilos['Normal']
    estilo_celda.fontSize = 7.5
    estilo_celda.alignment = 0
    estilo_celda.spaceBefore = 0
    estilo_celda.spaceAfter = 0

    estilo_cabecera = estilos['Normal']
    estilo_cabecera.fontSize = 8
    estilo_cabecera.textColor = colors.white
    estilo_cabecera.alignment = 1

    datos = [[Paragraph(col, estilo_cabecera) for col in columnas]]
    datos += [[Paragraph(str(celda or ''), estilo_celda) for celda in fila] for fila in filas]

    ancho_columna = doc.width / len(columnas)
    tabla = Table(datos, colWidths=[ancho_columna] * len(columnas), repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#9AA5B1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EAF1F8')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    generado = datetime.now().strftime('%d/%m/%Y %H:%M')
    contenido = [
        Paragraph(titulo, estilo_titulo),
        Paragraph(f"Generado el {generado}", estilo_subtitulo),
        Spacer(1, 4),
        tabla,
    ]
    doc.build(contenido)
    return buffer.getvalue()


def _build_xlsx(titulo, columnas, filas):
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='1F4E79')
        celda.alignment = Alignment(horizontal='center', vertical='center')

    for fila in filas:
        ws.append([str(celda or '') if not isinstance(celda, int) else celda for celda in fila])

    for idx, columna in enumerate(columnas, start=1):
        ancho = len(columna) + 4
        for fila in ws.iter_rows(min_col=idx, max_col=idx):
            for celda in fila:
                ancho = max(ancho, len(str(celda.value or '')) + 2)
        ws.column_dimensions[get_column_letter(idx)].width = min(ancho, 60)

    ws.freeze_panes = 'A2'
    wb.save(buffer)
    return buffer.getvalue()


def _responder_exportacion(titulo, columnas, filas, formato, nombre):
    formato = (formato or 'xlsx').lower()
    if formato == 'pdf':
        contenido = _build_pdf(titulo, columnas, filas)
        content_type = 'application/pdf'
        extension = 'pdf'
    else:
        contenido = _build_xlsx(titulo, columnas, filas)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        extension = 'xlsx'

    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    respuesta = HttpResponse(contenido, content_type=content_type)
    respuesta['Content-Disposition'] = f'attachment; filename="{nombre}_{fecha}.{extension}"'
    return respuesta


COLUMNAS_OBSERVACIONES = ['Evidencia', 'Indicador', 'Departamento', 'Periodo', 'Versión', 'Observador', 'Comentario', 'Fecha', 'N° Observaciones']
COLUMNAS_AUDITORIA = ['Usuario', 'Acción', 'Modelo', 'Registro ID', 'Descripción', 'Fecha']
COLUMNAS_USUARIOS = ['Usuario', 'Nombre', 'Correo', 'Rol', 'Departamento', 'Último acceso', 'Estado']


# ============================================================
# Reporte 5: Observaciones
# ============================================================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def observaciones(request):
    qs = _base_queryset_observaciones(request)
    filas_completas = _filas_observaciones(qs)

    evidencias_observadas = len(set(fila[0] for fila in filas_completas))
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'evidencia': fila[0],
            'indicador': fila[1],
            'departamento': fila[2],
            'periodo': fila[3],
            'version': fila[4],
            'observador': fila[5],
            'comentario': fila[6],
            'fecha': fila[7],
            'total_observaciones': fila[8],
        }
        for fila in filas
    ]

    return Response({
        'total_observaciones': total,
        'evidencias_observadas': evidencias_observadas,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def observaciones_exportar(request):
    qs = _base_queryset_observaciones(request)
    filas = _filas_observaciones(qs)
    return _responder_exportacion(
        'Reporte de Observaciones',
        COLUMNAS_OBSERVACIONES,
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_observaciones',
    )


# ============================================================
# Reporte 6: Auditoría
# ============================================================
def _base_queryset_auditoria(request):
    qs = Auditoria.objects.select_related('usuario')

    usuario = request.query_params.get('usuario')
    if usuario:
        qs = qs.filter(usuario_id=usuario)

    fecha_desde = request.query_params.get('fecha_desde')
    if fecha_desde:
        qs = qs.filter(fecha__date__gte=fecha_desde)

    fecha_hasta = request.query_params.get('fecha_hasta')
    if fecha_hasta:
        qs = qs.filter(fecha__date__lte=fecha_hasta)

    modelo = request.query_params.get('modelo')
    if modelo:
        qs = qs.filter(modelo__icontains=modelo)

    accion = request.query_params.get('accion')
    if accion:
        qs = qs.filter(accion__icontains=accion)

    return qs


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def auditoria(request):
    qs = _base_queryset_auditoria(request)
    filas_completas = _filas_auditoria(qs)
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'usuario_nombre': fila[0],
            'accion': fila[1],
            'modelo': fila[2],
            'registro_id': fila[3],
            'descripcion': fila[4],
            'fecha': fila[5],
        }
        for fila in filas
    ]

    return Response({
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def auditoria_exportar(request):
    qs = _base_queryset_auditoria(request)
    filas = _filas_auditoria(qs)
    return _responder_exportacion(
        'Reporte de Auditoría',
        COLUMNAS_AUDITORIA,
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_auditoria',
    )


# ============================================================
# Reporte 7: Usuarios
# ============================================================
def _base_queryset_usuarios(request):
    qs = Usuario.objects.all().prefetch_related('groups').select_related('perfilusuario__departamento')

    rol = request.query_params.get('rol')
    if rol:
        qs = qs.filter(groups__name__icontains=rol)

    departamento = request.query_params.get('departamento')
    if departamento:
        qs = qs.filter(perfilusuario__departamento_id=departamento)

    estado = request.query_params.get('estado')
    if estado:
        qs = qs.filter(is_active=(estado == 'activo'))

    return qs.order_by('username')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def usuarios(request):
    qs = _base_queryset_usuarios(request)
    filas_completas = _filas_usuarios(qs)
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'username': fila[0],
            'nombre': fila[1],
            'email': fila[2],
            'rol': fila[3],
            'departamento_nombre': fila[4],
            'ultimo_acceso': fila[5],
            'estado': fila[6],
        }
        for fila in filas
    ]

    activos = sum(1 for fila in filas_completas if fila[6] == 'Activo')
    return Response({
        'total': total,
        'activos': activos,
        'inactivos': total - activos,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def usuarios_exportar(request):
    qs = _base_queryset_usuarios(request)
    filas = _filas_usuarios(qs)
    return _responder_exportacion(
        'Reporte de Usuarios',
        COLUMNAS_USUARIOS,
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_usuarios',
    )
