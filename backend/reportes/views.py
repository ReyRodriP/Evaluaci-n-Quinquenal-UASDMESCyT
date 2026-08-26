"""
@file views.py
@brief Vistas de la app de reportes.
@details Define las vistas para la generación y exportación de reportes
del sistema de evaluación quinquenal, incluyendo reportes generales,
por facultad, por departamento, de evidencias, observaciones,
auditoría y usuarios.
"""

import io
import re
from datetime import datetime
from math import ceil

from django.contrib.auth import get_user_model
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.permissions import (
    PuedeVerReportes,
    PuedeVerReportesCompletos,
    departamentos_permitidos,
    facultades_permitidas,
)
from auditoria.models import Auditoria
from evaluation.models import Asignacion, Criterio, EstadoAsignacion, HistorialEstado, Periodo
from evidence.models import Evidencia, Observacion, VersionEvidencia
from organization.models import Departamento, Facultad, PerfilUsuario

Usuario = get_user_model()

FILTRO_POR_DEFECTO = {
    'fecha': 30,
}


def _formato_fecha(valor):
    if not valor:
        return ''
    return valor.strftime('%d/%m/%Y %H:%M')


def _formato_fecha_corta(valor):
    if not valor:
        return ''
    return valor.strftime('%d/%m/%Y')


def _base_queryset_observaciones(request):
    qs = Observacion.objects.filter(activo=True).select_related(
        'usuario',
        'version__evidencia__asignacion__indicador',
        'version__evidencia__asignacion__departamento',
        'version__evidencia__asignacion__periodo',
    )

    deptos_ids = departamentos_permitidos(request)
    if deptos_ids is not None:
        qs = qs.filter(version__evidencia__asignacion__departamento_id__in=deptos_ids)

    periodo = request.query_params.get('periodo')
    if periodo:
        qs = qs.filter(version__evidencia__asignacion__periodo_id=periodo)

    departamento = request.query_params.get('departamento')
    if departamento:
        qs = qs.filter(version__evidencia__asignacion__departamento_id=departamento)

    usuario = request.query_params.get('usuario')
    if usuario:
        qs = qs.filter(usuario_id=usuario)

    return qs


def _filas_observaciones(qs):
    conteos = dict(
        qs.values_list('version__evidencia_id')
        .annotate(total=Count('id'))
        .order_by()
    )

    filas = []
    for obs in qs:
        asignacion = obs.version.evidencia.asignacion
        filas.append([
            obs.version.evidencia.titulo,
            asignacion.indicador.nombre,
            asignacion.departamento.nombre,
            asignacion.periodo.nombre,
            obs.version.version,
            obs.usuario.username if obs.usuario else '',
            obs.comentario,
            _formato_fecha(obs.fecha_creacion),
            conteos.get(obs.version.evidencia_id, 0),
        ])

    return filas


def _filas_auditoria(qs):
    filas = []
    for registro in qs:
        filas.append([
            registro.usuario.username if registro.usuario else '',
            registro.accion,
            registro.modelo,
            registro.registro_id or '',
            registro.descripcion,
            _formato_fecha(registro.fecha),
        ])
    return filas


def _filas_usuarios(qs):
    filas = []
    for usuario in qs:
        perfil = getattr(usuario, 'perfilusuario', None)
        filas.append([
            usuario.username,
            f"{usuario.first_name} {usuario.last_name}".strip(),
            usuario.email,
            (usuario.groups.first().name if usuario.groups.exists() else ''),
            perfil.departamento.nombre if perfil and perfil.departamento else '',
            _formato_fecha(usuario.last_login),
            'Activo' if usuario.is_active else 'Inactivo',
        ])
    return filas


def _paginar(request, filas):
    try:
        page = max(int(request.query_params.get('page', 1)), 1)
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(max(int(request.query_params.get('page_size', 10)), 1), 100)
    except (TypeError, ValueError):
        page_size = 10

    total = len(filas)
    inicio = (page - 1) * page_size
    return filas[inicio:inicio + page_size], total, page, page_size, ceil(total / page_size) if total else 0


def _build_pdf(titulo, columnas, filas):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=titulo,
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('titulo_reporte', parent=estilos['Title'], fontSize=14, spaceAfter=2)

    estilo_subtitulo = ParagraphStyle('subtitulo_reporte', parent=estilos['Normal'], fontSize=9, alignment=1, spaceAfter=10)

    estilo_celda = ParagraphStyle('celda_reporte', parent=estilos['Normal'], fontSize=7.5, alignment=0, spaceBefore=0, spaceAfter=0, textColor=colors.black)

    estilo_cabecera = ParagraphStyle('cabecera_reporte', parent=estilos['Normal'], fontSize=8, textColor=colors.white, alignment=1)

    datos = [[Paragraph(col, estilo_cabecera) for col in columnas]]
    datos += [[Paragraph(str(celda or ''), estilo_celda) for celda in fila] for fila in filas]

    ancho_columna = doc.width / len(columnas)
    tabla = Table(datos, colWidths=[ancho_columna] * len(columnas), repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#9AA5B1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EAF1F8')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    generado = datetime.now().strftime('%d/%m/%Y %H:%M')
    contenido = [
        Paragraph(titulo, estilo_titulo),
        Paragraph(f"Generado el {generado}", estilo_subtitulo),
        Spacer(1, 4),
        tabla,
    ]
    doc.build(contenido)
    return buffer.getvalue()


def _build_xlsx(titulo, columnas, filas):
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    nombre_hoja = re.sub(r'[\[\]:*?/\\]', '', titulo)[:31] or 'Reporte'
    ws.title = nombre_hoja

    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='1F4E79')
        celda.alignment = Alignment(horizontal='center', vertical='center')

    for fila in filas:
        ws.append([str(celda or '') if not isinstance(celda, int) else celda for celda in fila])

    for idx, columna in enumerate(columnas, start=1):
        ancho = len(columna) + 4
        for fila in ws.iter_rows(min_col=idx, max_col=idx):
            for celda in fila:
                ancho = max(ancho, len(str(celda.value or '')) + 2)
        ws.column_dimensions[get_column_letter(idx)].width = min(ancho, 60)

    ws.freeze_panes = 'A2'
    wb.save(buffer)
    return buffer.getvalue()


def _responder_exportacion(titulo, columnas, filas, formato, nombre):
    formato = (formato or 'xlsx').lower()
    if formato == 'pdf':
        contenido = _build_pdf(titulo, columnas, filas)
        content_type = 'application/pdf'
        extension = 'pdf'
    else:
        contenido = _build_xlsx(titulo, columnas, filas)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        extension = 'xlsx'

    fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
    respuesta = HttpResponse(contenido, content_type=content_type)
    respuesta['Content-Disposition'] = f'attachment; filename="{nombre}_{fecha}.{extension}"'
    return respuesta


COLUMNAS_OBSERVACIONES = ['Evidencia', 'Indicador', 'Departamento', 'Periodo', 'Versión', 'Observador', 'Comentario', 'Fecha', 'N° Observaciones']
COLUMNAS_AUDITORIA = ['Usuario', 'Acción', 'Modelo', 'Registro ID', 'Descripción', 'Fecha']
COLUMNAS_USUARIOS = ['Usuario', 'Nombre', 'Correo', 'Rol', 'Departamento', 'Último acceso', 'Estado']


# ============================================================
# Reporte 1: General del periodo
# ============================================================
def _data_general(request):
    periodo_id = request.query_params.get('periodo')
    if periodo_id:
        periodo = get_object_or_404(Periodo, pk=periodo_id)
    else:
        periodo = Periodo.objects.order_by('-id').first()

    deptos_ids = departamentos_permitidos(request)

    qs = Asignacion.objects.all()
    if deptos_ids is not None:
        qs = qs.filter(departamento_id__in=deptos_ids)
    if periodo:
        qs = qs.filter(periodo=periodo)

    total_asignaciones = qs.count()

    return {
        'periodo': periodo.nombre if periodo else None,
        'total_departamentos': (
            qs.values('departamento_id').distinct().count()
            if total_asignaciones else (
                len(deptos_ids) if deptos_ids is not None else Departamento.objects.count()
            )
        ),
        'total_indicadores': qs.values('indicador_id').distinct().count(),
        'total_asignaciones': total_asignaciones,
        'pendientes': qs.filter(estado=EstadoAsignacion.PENDIENTE).count(),
        'en_revision': qs.filter(estado=EstadoAsignacion.EN_PROGRESO).count(),
        'observadas': qs.filter(estado=EstadoAsignacion.OBSERVADA).count(),
        'aprobadas': qs.filter(estado=EstadoAsignacion.APROBADO).count(),
        'rechazadas': qs.filter(estado=EstadoAsignacion.RECHAZADO).count(),
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def general(request):
    """@brief Retorna datos del reporte general del período.
    @details Calcula estadísticas generales de asignaciones para el
    período actual o uno específico.
    @param request Request HTTP autenticada.
    @return Response con datos del reporte general.
    """
    return Response(_data_general(request))


def _filas_general(data):
    return [
        ['Periodo', data['periodo'] or ''],
        ['Total departamentos', data['total_departamentos']],
        ['Total indicadores', data['total_indicadores']],
        ['Total asignaciones', data['total_asignaciones']],
        ['Pendientes', data['pendientes']],
        ['En revisión', data['en_revision']],
        ['Observadas', data['observadas']],
        ['Aprobadas', data['aprobadas']],
        ['Rechazadas', data['rechazadas']],
    ]


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def general_exportar(request):
    """@brief Exporta el reporte general en PDF o XLSX.
    @param request Request HTTP autenticada.
    @return HttpResponse con el archivo generado.
    """
    data = _data_general(request)
    return _responder_exportacion(
        'Reporte General del Periodo',
        ['Métrica', 'Valor'],
        _filas_general(data),
        request.query_params.get('formato', 'xlsx'),
        'reporte_general',
    )


# ============================================================
# Reporte 2: Por facultad
# ============================================================
def _base_departamentos_visibles(request):
    deptos_ids = departamentos_permitidos(request)
    qs = Departamento.objects.all()
    if deptos_ids is not None:
        qs = qs.filter(pk__in=deptos_ids)
    return qs


def _data_por_facultad(request, pk):
    facultad = get_object_or_404(Facultad, pk=pk)

    facultades_ids = facultades_permitidas(request)
    if facultades_ids is not None and facultad.pk not in facultades_ids:
        return {'facultad': facultad.nombre, 'denegado': True, 'departamentos': [], 'pendientes': 0, 'aprobadas': 0}

    deptos_ids = departamentos_permitidos(request)
    departamentos = Departamento.objects.filter(facultad=facultad)
    if deptos_ids is not None:
        departamentos = departamentos.filter(pk__in=deptos_ids)

    asignaciones = Asignacion.objects.filter(departamento__in=departamentos)

    data = {
        'facultad': facultad.nombre,
        'departamentos': [],
        'pendientes': asignaciones.filter(estado=EstadoAsignacion.PENDIENTE).count(),
        'aprobadas': asignaciones.filter(estado=EstadoAsignacion.APROBADO).count(),
    }

    for departamento in departamentos:
        asignaciones_dep = Asignacion.objects.filter(departamento=departamento)
        data['departamentos'].append({
            'id': departamento.pk,
            'nombre': departamento.nombre,
            'total_asignaciones': asignaciones_dep.count(),
            'evidencias': Evidencia.objects.filter(asignacion__departamento=departamento).count(),
            'pendientes': asignaciones_dep.filter(estado=EstadoAsignacion.PENDIENTE).count(),
            'aprobadas': asignaciones_dep.filter(estado=EstadoAsignacion.APROBADO).count(),
        })

    return data


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def por_facultad(request, pk=None):
    """@brief Retorna datos del reporte por facultad.
    @param request Request HTTP autenticada.
    @param pk Identificador de la facultad.
    @return Response con datos del reporte o error 403/404.
    """
    data = _data_por_facultad(request, pk)
    if data.get('denegado'):
        return Response({'detail': 'No autorizado'}, status=403)
    return Response(data)


def _filas_facultad(data):
    filas = []
    for dept in data['departamentos']:
        filas.append([
            dept['nombre'],
            dept['total_asignaciones'],
            dept['evidencias'],
            dept['pendientes'],
            dept['aprobadas'],
        ])
    return filas


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def por_facultad_exportar(request, pk=None):
    """@brief Exporta el reporte por facultad en PDF o XLSX.
    @param request Request HTTP autenticada.
    @param pk Identificador de la facultad.
    @return HttpResponse con el archivo generado.
    """
    data = _data_por_facultad(request, pk)
    if data.get('denegado'):
        return Response({'detail': 'No autorizado'}, status=403)
    return _responder_exportacion(
        f"Reporte por Facultad: {data['facultad']}",
        ['Departamento', 'Total asignaciones', 'Evidencias', 'Pendientes', 'Aprobadas'],
        _filas_facultad(data),
        request.query_params.get('formato', 'xlsx'),
        f"reporte_facultad_{pk}",
    )


# ============================================================
# Reporte 3: Por departamento
# ============================================================
def _responsable_departamento(departamento):
    perfiles = PerfilUsuario.objects.filter(departamento=departamento).select_related('usuario').prefetch_related('usuario__groups')
    for perfil in perfiles:
        if perfil.usuario.groups.filter(name='Responsable Departamental').exists():
            return perfil.usuario.username
    if perfiles.exists():
        return perfiles.first().usuario.username
    return ''


def _data_por_departamento(request, pk):
    departamento = get_object_or_404(Departamento, pk=pk)

    deptos_ids = departamentos_permitidos(request)
    if deptos_ids is not None and departamento.pk not in deptos_ids:
        return {'departamento': departamento.nombre, 'denegado': True, 'total': 0, 'rows': []}

    asignaciones = Asignacion.objects.filter(departamento=departamento).select_related('indicador')

    filas = []
    for asignacion in asignaciones:
        ultima_version = VersionEvidencia.objects.filter(
            evidencia__asignacion=asignacion
        ).order_by('-fecha_subida').first()

        ultimo_historial = HistorialEstado.objects.filter(
            asignacion=asignacion
        ).order_by('-fecha').first()

        filas.append({
            'indicador': asignacion.indicador.nombre,
            'estado': asignacion.estado,
            'fecha_modificacion': (
                _formato_fecha(ultimo_historial.fecha) if ultimo_historial else ''
            ),
            'responsable': _responsable_departamento(departamento),
            'ultima_version': ultima_version.version if ultima_version else None,
        })

    filas, total, page, page_size, total_pages = _paginar(request, filas)

    return {
        'departamento': departamento.nombre,
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas,
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def por_departamento(request, pk=None):
    """@brief Retorna datos del reporte por departamento.
    @param request Request HTTP autenticada.
    @param pk Identificador del departamento.
    @return Response con datos del reporte o error 403/404.
    """
    data = _data_por_departamento(request, pk)
    if data.get('denegado'):
        return Response({'detail': 'No autorizado'}, status=403)
    return Response(data)


def _filas_departamento(data):
    return [
        [
            fila['indicador'],
            fila['estado'],
            fila['fecha_modificacion'],
            fila['responsable'],
            fila['ultima_version'],
        ]
        for fila in data['rows']
    ]


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def por_departamento_exportar(request, pk=None):
    """@brief Exporta el reporte por departamento en PDF o XLSX.
    @param request Request HTTP autenticada.
    @param pk Identificador del departamento.
    @return HttpResponse con el archivo generado.
    """
    data = _data_por_departamento(request, pk)
    if data.get('denegado'):
        return Response({'detail': 'No autorizado'}, status=403)
    return _responder_exportacion(
        f"Reporte por Departamento: {data['departamento']}",
        ['Indicador', 'Estado', 'Fecha modificación', 'Responsable', 'Última versión'],
        _filas_departamento(data),
        request.query_params.get('formato', 'xlsx'),
        f"reporte_departamento_{pk}",
    )


# ============================================================
# Reporte 4: Evidencias
# ============================================================
def _base_queryset_evidencias(request):
    qs = Evidencia.objects.select_related(
        'asignacion__indicador__criterio',
        'asignacion__departamento',
        'asignacion__periodo',
    )

    deptos_ids = departamentos_permitidos(request)
    if deptos_ids is not None:
        qs = qs.filter(asignacion__departamento_id__in=deptos_ids)

    estado = request.query_params.get('estado')
    if estado:
        qs = qs.filter(asignacion__estado=estado)

    departamento = request.query_params.get('departamento')
    if departamento:
        qs = qs.filter(asignacion__departamento_id=departamento)

    periodo = request.query_params.get('periodo')
    if periodo:
        qs = qs.filter(asignacion__periodo_id=periodo)

    criterio = request.query_params.get('criterio')
    if criterio:
        qs = qs.filter(asignacion__indicador__criterio_id=criterio)

    return qs


def _filas_evidencias(qs):
    filas = []
    for evidencia in qs:
        ultima_version = evidencia.versiones.order_by('-fecha_subida').first()
        asignacion = evidencia.asignacion
        filas.append([
            asignacion.indicador.nombre,
            evidencia.titulo,
            asignacion.departamento.nombre,
            asignacion.indicador.criterio.nombre,
            asignacion.periodo.nombre,
            asignacion.estado,
            _formato_fecha(ultima_version.fecha_subida) if ultima_version else '',
        ])
    return filas


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def evidencias(request):
    """@brief Retorna datos del reporte de evidencias.
    @details Lista evidencias con paginación y filtros por estado,
    departamento, período y criterio.
    @param request Request HTTP autenticada.
    @return Response con evidencias paginadas.
    """
    qs = _base_queryset_evidencias(request)
    filas_completas = _filas_evidencias(qs)
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'indicador': fila[0],
            'evidencia': fila[1],
            'departamento': fila[2],
            'criterio': fila[3],
            'periodo': fila[4],
            'estado': fila[5],
            'fecha': fila[6],
        }
        for fila in filas
    ]

    return Response({
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def evidencias_exportar(request):
    """@brief Exporta el reporte de evidencias en PDF o XLSX.
    @param request Request HTTP autenticada.
    @return HttpResponse con el archivo generado.
    """
    qs = _base_queryset_evidencias(request)
    filas = _filas_evidencias(qs)
    return _responder_exportacion(
        'Reporte de Evidencias',
        ['Indicador', 'Evidencia', 'Departamento', 'Criterio', 'Periodo', 'Estado', 'Fecha'],
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_evidencias',
    )


# ============================================================
# Reporte 5: Observaciones
# ============================================================
@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def observaciones(request):
    """@brief Retorna datos del reporte de observaciones.
    @details Lista observaciones de evidencias con paginación
    y filtros por período, departamento y usuario.
    @param request Request HTTP autenticada.
    @return Response con observaciones paginadas.
    """
    qs = _base_queryset_observaciones(request)
    filas_completas = _filas_observaciones(qs)

    evidencias_observadas = len(set(fila[0] for fila in filas_completas))
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'evidencia': fila[0],
            'indicador': fila[1],
            'departamento': fila[2],
            'periodo': fila[3],
            'version': fila[4],
            'observador': fila[5],
            'comentario': fila[6],
            'fecha': fila[7],
            'total_observaciones': fila[8],
        }
        for fila in filas
    ]

    return Response({
        'total_observaciones': total,
        'evidencias_observadas': evidencias_observadas,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportes])
def observaciones_exportar(request):
    """@brief Exporta el reporte de observaciones en PDF o XLSX.
    @param request Request HTTP autenticada.
    @return HttpResponse con el archivo generado.
    """
    qs = _base_queryset_observaciones(request)
    filas = _filas_observaciones(qs)
    return _responder_exportacion(
        'Reporte de Observaciones',
        COLUMNAS_OBSERVACIONES,
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_observaciones',
    )


# ============================================================
# Reporte 6: Auditoría
# ============================================================
def _base_queryset_auditoria(request):
    qs = Auditoria.objects.select_related('usuario')

    usuario = request.query_params.get('usuario')
    if usuario:
        qs = qs.filter(usuario_id=usuario)

    fecha_desde = request.query_params.get('fecha_desde')
    if fecha_desde:
        qs = qs.filter(fecha__date__gte=fecha_desde)

    fecha_hasta = request.query_params.get('fecha_hasta')
    if fecha_hasta:
        qs = qs.filter(fecha__date__lte=fecha_hasta)

    modelo = request.query_params.get('modelo')
    if modelo:
        qs = qs.filter(modelo__icontains=modelo)

    accion = request.query_params.get('accion')
    if accion:
        qs = qs.filter(accion__icontains=accion)

    return qs


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportesCompletos])
def auditoria(request):
    """@brief Retorna datos del reporte de auditoría.
    @details Lista registros de auditoría con paginación y filtros
    por usuario, fechas, modelo y acción.
    @param request Request HTTP autenticada.
    @return Response con registros de auditoría paginados.
    """
    qs = _base_queryset_auditoria(request)
    filas_completas = _filas_auditoria(qs)
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'usuario_nombre': fila[0],
            'accion': fila[1],
            'modelo': fila[2],
            'registro_id': fila[3],
            'descripcion': fila[4],
            'fecha': fila[5],
        }
        for fila in filas
    ]

    return Response({
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportesCompletos])
def auditoria_exportar(request):
    """@brief Exporta el reporte de auditoría en PDF o XLSX.
    @param request Request HTTP autenticada.
    @return HttpResponse con el archivo generado.
    """
    qs = _base_queryset_auditoria(request)
    filas = _filas_auditoria(qs)
    return _responder_exportacion(
        'Reporte de Auditoría',
        COLUMNAS_AUDITORIA,
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_auditoria',
    )


# ============================================================
# Reporte 7: Usuarios
# ============================================================
def _base_queryset_usuarios(request):
    qs = Usuario.objects.all().prefetch_related('groups').select_related('perfilusuario__departamento')

    rol = request.query_params.get('rol')
    if rol:
        qs = qs.filter(groups__name__icontains=rol)

    departamento = request.query_params.get('departamento')
    if departamento:
        qs = qs.filter(perfilusuario__departamento_id=departamento)

    estado = request.query_params.get('estado')
    if estado:
        qs = qs.filter(is_active=(estado == 'activo'))

    return qs.order_by('username')


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportesCompletos])
def usuarios(request):
    """@brief Retorna datos del reporte de usuarios.
    @details Lista usuarios con paginación y filtros por rol,
    departamento y estado.
    @param request Request HTTP autenticada.
    @return Response con usuarios paginados.
    """
    qs = _base_queryset_usuarios(request)
    filas_completas = _filas_usuarios(qs)
    filas, total, page, page_size, total_pages = _paginar(request, filas_completas)

    filas_json = [
        {
            'username': fila[0],
            'nombre': fila[1],
            'email': fila[2],
            'rol': fila[3],
            'departamento_nombre': fila[4],
            'ultimo_acceso': fila[5],
            'estado': fila[6],
        }
        for fila in filas
    ]

    activos = sum(1 for fila in filas_completas if fila[6] == 'Activo')
    return Response({
        'total': total,
        'activos': activos,
        'inactivos': total - activos,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
        'rows': filas_json,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportesCompletos])
def usuarios_exportar(request):
    """@brief Exporta el reporte de usuarios en PDF o XLSX.
    @param request Request HTTP autenticada.
    @return HttpResponse con el archivo generado.
    """
    qs = _base_queryset_usuarios(request)
    filas = _filas_usuarios(qs)
    return _responder_exportacion(
        'Reporte de Usuarios',
        COLUMNAS_USUARIOS,
        filas,
        request.query_params.get('formato', 'xlsx'),
        'reporte_usuarios',
    )
